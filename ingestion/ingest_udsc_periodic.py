"""
ingest_udsc_periodic.py

Validates and uploads the UdSC first-half reports used for a like-for-like
comparison of January-June 2025 and January-June 2026.

Official sources:
https://www.gov.pl/web/udsc/2025r
https://www.gov.pl/web/udsc/raporty2026
"""

import hashlib
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from azure.core.exceptions import ResourceExistsError
from azure.storage.filedatalake import DataLakeServiceClient
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()

STORAGE_ACCOUNT_NAME = os.getenv("AZURE_STORAGE_ACCOUNT_NAME")
STORAGE_ACCOUNT_KEY = os.getenv("AZURE_STORAGE_ACCOUNT_KEY")

RAW_CONTAINER = "raw"
LOCAL_DIRECTORY = Path("data/raw/udsc/periodic")
AZURE_BASE_DIRECTORY = "udsc/periodic"
EXPECTED_YEARS = {2025, 2026}
EXPECTED_SHEETS = {"Arkusz9", "Arkusz11", "Arkusz18"}

SOURCE_URLS = {
    2025: "https://www.gov.pl/web/udsc/2025r",
    2026: "https://www.gov.pl/web/udsc/raporty2026",
}


def validate_credentials():
    """Checks that the Azure Storage credentials exist."""

    if not STORAGE_ACCOUNT_NAME or not STORAGE_ACCOUNT_KEY:
        raise ValueError(
            "Missing AZURE_STORAGE_ACCOUNT_NAME or "
            "AZURE_STORAGE_ACCOUNT_KEY in .env"
        )


def calculate_sha256(filepath):
    """Calculates a checksum for source-file traceability."""

    sha256 = hashlib.sha256()
    with filepath.open("rb") as file:
        for chunk in iter(lambda: file.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def parse_polish_date(value):
    """Converts the date strings stored in Arkusz18."""

    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(str(value).strip(), "%d.%m.%Y").date()


def inspect_workbook(filepath):
    """Validates the workbook and returns its reporting-period metadata."""

    workbook = load_workbook(filepath, read_only=True, data_only=True)
    missing_sheets = EXPECTED_SHEETS - set(workbook.sheetnames)
    if missing_sheets:
        workbook.close()
        raise ValueError(
            f"{filepath.name} is missing sheets: {sorted(missing_sheets)}"
        )

    period_sheet = workbook["Arkusz18"]
    month_start = parse_polish_date(period_sheet.cell(2, 1).value)
    period_end = parse_polish_date(period_sheet.cell(2, 2).value)
    period_start = parse_polish_date(period_sheet.cell(2, 3).value)
    sheet_names = workbook.sheetnames
    workbook.close()

    if period_start.month != 1 or period_start.day != 1:
        raise ValueError(f"Unexpected period start in {filepath.name}")
    if period_end.month != 6 or period_end.day != 30:
        raise ValueError(f"The report is not a first-half report: {filepath.name}")
    if month_start.month != 6 or month_start.day != 1:
        raise ValueError(f"Unexpected monthly boundary in {filepath.name}")
    if len({period_start.year, month_start.year, period_end.year}) != 1:
        raise ValueError(f"Mixed reporting years in {filepath.name}")

    return {
        "year": period_end.year,
        "period_start_date": period_start.isoformat(),
        "period_end_date": period_end.isoformat(),
        "period_type": "half_year",
        "half_number": 1,
        "is_complete_year": False,
        "sheet_names": sheet_names,
    }


def discover_periodic_files():
    """Finds exactly one valid first-half workbook for 2025 and 2026."""

    if not LOCAL_DIRECTORY.exists():
        raise FileNotFoundError(
            f"Local directory does not exist: {LOCAL_DIRECTORY}"
        )

    files_by_year = {}
    metadata_by_year = {}

    for filepath in LOCAL_DIRECTORY.rglob("*.xlsx"):
        if filepath.name.startswith("~$"):
            continue

        metadata = inspect_workbook(filepath)
        year = metadata["year"]

        if year not in EXPECTED_YEARS:
            print(f"Skipping unexpected year {year}: {filepath.name}")
            continue
        if year in files_by_year:
            raise ValueError(f"More than one first-half report found for {year}")

        files_by_year[year] = filepath
        metadata_by_year[year] = metadata

    missing_years = EXPECTED_YEARS - set(files_by_year)
    if missing_years:
        raise ValueError(
            f"Missing first-half reports for: {sorted(missing_years)}"
        )

    return files_by_year, metadata_by_year


def get_file_system_client():
    """Creates the connection to the Azure raw container."""

    validate_credentials()
    service_client = DataLakeServiceClient(
        account_url=(
            f"https://{STORAGE_ACCOUNT_NAME}.dfs.core.windows.net"
        ),
        credential=STORAGE_ACCOUNT_KEY,
    )
    return service_client.get_file_system_client(RAW_CONTAINER)


def ensure_directory(file_system_client, directory_path):
    """Creates every level of an Azure Data Lake directory."""

    current_path = ""
    for part in directory_path.split("/"):
        current_path = f"{current_path}/{part}" if current_path else part
        try:
            file_system_client.get_directory_client(
                current_path
            ).create_directory()
        except ResourceExistsError:
            pass


def upload_bytes(file_system_client, azure_path, data):
    """Uploads bytes to a specific Azure Data Lake path."""

    parent_directory = str(Path(azure_path).parent).replace("\\", "/")
    ensure_directory(file_system_client, parent_directory)
    file_system_client.get_file_client(azure_path).upload_data(
        data,
        overwrite=True,
    )


def upload_reports(files_by_year, metadata_by_year):
    """Uploads the two reports and returns their manifest entries."""

    file_system_client = get_file_system_client()
    manifest_files = []

    for year in sorted(files_by_year):
        filepath = files_by_year[year]
        metadata = metadata_by_year[year]
        azure_path = (
            f"{AZURE_BASE_DIRECTORY}/{year}/first_half/{filepath.name}"
        )

        print(f"Uploading {year} H1 to raw/{azure_path}")
        upload_bytes(file_system_client, azure_path, filepath.read_bytes())

        manifest_files.append(
            {
                **{key: value for key, value in metadata.items() if key != "sheet_names"},
                "original_filename": filepath.name,
                "source_url": SOURCE_URLS[year],
                "azure_path": f"{RAW_CONTAINER}/{azure_path}",
                "size_bytes": filepath.stat().st_size,
                "sha256": calculate_sha256(filepath),
                "sheet_count": len(metadata["sheet_names"]),
                "sheet_names": metadata["sheet_names"],
            }
        )
        print(f"Uploaded successfully: {year} H1")

    return file_system_client, manifest_files


def upload_manifest(file_system_client, manifest_files):
    """Uploads the periodic-report ingestion manifest."""

    manifest = {
        "source_system": "UdSC",
        "source_name": "Office for Foreigners",
        "dataset": "First-half residence proceedings",
        "period_type": "half_year",
        "years": sorted(EXPECTED_YEARS),
        "is_complete_year": False,
        "ingested_at_utc": datetime.now(timezone.utc).isoformat(),
        "files": manifest_files,
    }
    manifest_path = f"{AZURE_BASE_DIRECTORY}/_manifest.json"
    upload_bytes(
        file_system_client,
        manifest_path,
        json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"),
    )
    print(f"Uploaded manifest to raw/{manifest_path}")


def main():
    """Runs the complete first-half UdSC ingestion."""

    print("Discovering UdSC first-half reports...")
    files_by_year, metadata_by_year = discover_periodic_files()
    print(f"Years discovered: {sorted(files_by_year)}")
    file_system_client, manifest_files = upload_reports(
        files_by_year,
        metadata_by_year,
    )
    upload_manifest(file_system_client, manifest_files)
    print("Done. UdSC first-half ingestion completed successfully.")


if __name__ == "__main__":
    main()
