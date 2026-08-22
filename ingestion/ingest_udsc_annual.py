"""
ingest_udsc_annual.py

Validates and uploads annual UdSC Excel reports to the Azure Data Lake
raw container, preserving the original source files.

Official source:
https://www.gov.pl/web/udsc/zestawienia-roczne
"""

import hashlib
import io
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from azure.storage.filedatalake import DataLakeServiceClient
from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()

STORAGE_ACCOUNT_NAME = os.getenv("AZURE_STORAGE_ACCOUNT_NAME")
STORAGE_ACCOUNT_KEY = os.getenv("AZURE_STORAGE_ACCOUNT_KEY")

RAW_CONTAINER = "raw"
LOCAL_DIRECTORY = Path("data/raw/udsc/annual")
AZURE_BASE_DIRECTORY = "udsc/annual"

EXPECTED_YEARS = {2021, 2022, 2023, 2024, 2025}

EXPECTED_SHEETS = {
    "WNIOSKI_OCHRONA",
    "DECYZJE-OCHR",
    "ODWOŁANIA",
    "DECYZJE-RADA",
    "AZYL",
    "WIZY",
    "ZAPROSZENIA",
    "POB.STAŁY-WNIOSKI",
    "POB.STAŁY-DECYZJE",
    "REZYDENT-WNI",
    "REZYDENT-DEC",
    "POB.CZASOWY-WNIOSKI",
    "POB.CZASOWY-DECYZJE",
    "UNIA EUROPEJSKA",
    "RODZINY UE",
    "WIELKA BRYTANIA",
    "RODZINY WB",
    "ZOBOWIĄZANIA",
    "ODMOWA",
    "POBYT TOLEROWANY",
    "POBYT HUMANITARNY",
    "ODWOŁANIA - LEGALIZACJA",
    "KARTY POBYTU",
}

SOURCE_URL = "https://www.gov.pl/web/udsc/zestawienia-roczne"


def validate_credentials():
    """Checks that the Azure Storage credentials exist."""

    if not STORAGE_ACCOUNT_NAME or not STORAGE_ACCOUNT_KEY:
        raise ValueError(
            "Missing AZURE_STORAGE_ACCOUNT_NAME or "
            "AZURE_STORAGE_ACCOUNT_KEY in .env"
        )


def extract_year(filename):
    """Extracts a four-digit year from the source filename."""

    match = re.search(r"20\d{2}", filename)

    if not match:
        raise ValueError(f"Could not identify year from: {filename}")

    return int(match.group())


def calculate_sha256(filepath):
    """Calculates a checksum to identify the exact source file."""

    sha256 = hashlib.sha256()

    with filepath.open("rb") as file:
        for chunk in iter(lambda: file.read(8192), b""):
            sha256.update(chunk)

    return sha256.hexdigest()


def validate_workbook(filepath):
    """
    Checks that the workbook contains the expected UdSC sheets.

    Returns the list of sheet names.
    """

    workbook = load_workbook(
        filepath,
        read_only=True,
        data_only=False,
    )

    sheet_names = set(workbook.sheetnames)
    workbook.close()

    missing_sheets = EXPECTED_SHEETS - sheet_names

    if missing_sheets:
        raise ValueError(
            f"{filepath.name} is missing sheets: "
            f"{sorted(missing_sheets)}"
        )

    return sorted(sheet_names)


def discover_annual_files():
    """
    Finds the annual Excel files and validates that there is exactly
    one file for every expected year.
    """

    if not LOCAL_DIRECTORY.exists():
        raise FileNotFoundError(
            f"Local directory does not exist: {LOCAL_DIRECTORY}"
        )

    files_by_year = {}

    for filepath in LOCAL_DIRECTORY.glob("*.xlsx"):
        if filepath.name.startswith("~$"):
            continue

        year = extract_year(filepath.name)

        if year not in EXPECTED_YEARS:
            print(f"Skipping unexpected year {year}: {filepath.name}")
            continue

        if year in files_by_year:
            raise ValueError(
                f"More than one Excel file found for {year}"
            )

        files_by_year[year] = filepath

    missing_years = EXPECTED_YEARS - set(files_by_year)

    if missing_years:
        raise ValueError(
            f"Missing annual reports for: {sorted(missing_years)}"
        )

    return files_by_year


def get_file_system_client():
    """Creates the connection to the Azure raw container."""

    account_url = (
        f"https://{STORAGE_ACCOUNT_NAME}.dfs.core.windows.net"
    )

    service_client = DataLakeServiceClient(
        account_url=account_url,
        credential=STORAGE_ACCOUNT_KEY,
    )

    return service_client.get_file_system_client(
        file_system=RAW_CONTAINER
    )


def upload_bytes(file_system_client, azure_path, data):
    """Uploads bytes to a specific path in Azure Data Lake."""

    file_client = file_system_client.get_file_client(azure_path)
    file_client.upload_data(data, overwrite=True)


def upload_annual_files(files_by_year):
    """
    Validates and uploads every annual report.

    Returns metadata that will be stored in the ingestion manifest.
    """

    file_system_client = get_file_system_client()
    manifest_files = []

    for year in sorted(files_by_year):
        filepath = files_by_year[year]

        print(f"\nValidating {year}: {filepath.name}")
        sheet_names = validate_workbook(filepath)

        azure_path = (
            f"{AZURE_BASE_DIRECTORY}/{year}/{filepath.name}"
        )

        print(f"Uploading to raw/{azure_path}")

        file_bytes = filepath.read_bytes()

        upload_bytes(
            file_system_client=file_system_client,
            azure_path=azure_path,
            data=file_bytes,
        )

        manifest_files.append(
            {
                "year": year,
                "original_filename": filepath.name,
                "azure_path": f"{RAW_CONTAINER}/{azure_path}",
                "size_bytes": filepath.stat().st_size,
                "sha256": calculate_sha256(filepath),
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
            }
        )

        print(f"Uploaded successfully: {year}")

    return file_system_client, manifest_files


def upload_manifest(file_system_client, manifest_files):
    """Creates and uploads the ingestion manifest."""

    manifest = {
        "source_system": "UdSC",
        "source_name": "Office for Foreigners",
        "source_url": SOURCE_URL,
        "dataset": "Annual proceedings involving foreigners",
        "period_type": "annual",
        "years": sorted(EXPECTED_YEARS),
        "ingested_at_utc": datetime.now(timezone.utc).isoformat(),
        "files": manifest_files,
    }

    manifest_bytes = json.dumps(
        manifest,
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")

    manifest_path = f"{AZURE_BASE_DIRECTORY}/_manifest.json"

    upload_bytes(
        file_system_client=file_system_client,
        azure_path=manifest_path,
        data=manifest_bytes,
    )

    print(f"\nUploaded manifest to raw/{manifest_path}")


def main():
    """Runs the complete UdSC annual ingestion process."""

    validate_credentials()

    print("Discovering UdSC annual reports...")
    files_by_year = discover_annual_files()

    print(f"Reports discovered: {len(files_by_year)}")
    print(f"Years: {sorted(files_by_year)}")

    file_system_client, manifest_files = upload_annual_files(
        files_by_year
    )

    upload_manifest(
        file_system_client=file_system_client,
        manifest_files=manifest_files,
    )

    print("\nDone. UdSC annual ingestion completed successfully.")


if __name__ == "__main__":
    main()